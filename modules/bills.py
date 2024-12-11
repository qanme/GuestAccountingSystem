import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font
from utils import setup_window, execute_query
from config import DB_PATH


# Загрузка данных счетов из базы данных
def load_bills(tree, search_query=""):
    try:
        query = '''
        SELECT 
            b.bill_id,
            b.booking_id,
            g.guest_id || ' - ' || g.last_name || ' ' || g.first_name AS guest_info,
            bk.room_number || ' - ' || r.room_type AS room_info,
            b.total_amount,
            b.payment_status
        FROM Bills b
        JOIN Bookings bk ON b.booking_id = bk.booking_id
        JOIN Guests g ON bk.guest_id = g.guest_id
        JOIN Rooms r ON bk.room_number = r.room_number
        WHERE 
            g.guest_id || ' - ' || g.last_name || ' ' || g.first_name LIKE ?
            OR bk.room_number || ' - ' || r.room_type LIKE ?
            OR CAST(b.bill_id AS TEXT) LIKE ?
            OR CAST(b.total_amount AS TEXT) LIKE ?
            OR b.payment_status LIKE ?
        ORDER BY b.bill_id DESC
        '''
        params = ('%' + search_query + '%',) * 5

        rows = execute_query(DB_PATH, query, params)

        # Очищаем дерево
        for row in tree.get_children():
            tree.delete(row)

        # Заполняем дерево новыми данными
        for row in rows:
            tree.insert("", "end", values=row)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить данные счетов: {e}")


# Удаление счета
def delete_bill(tree):
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Предупреждение", "Выберите счет для удаления!")
        return

    bill_id = tree.item(selected_item)["values"][0]

    try:
        # Подтверждение удаления счета
        confirmation = messagebox.askyesno("Подтверждение удаления", "Вы уверены, что хотите удалить этот счет?")
        if confirmation:
            query = "DELETE FROM Bills WHERE bill_id = ?"
            execute_query(DB_PATH, query, (bill_id,))
            messagebox.showinfo("Успех", "Счет удалён!")
            load_bills(tree)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось удалить счет: {e}")


# Изменение статуса оплаты
def change_payment_status(tree):
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Предупреждение", "Выберите счет для изменения статуса оплаты!")
        return

    bill_id = tree.item(selected_item)["values"][0]
    current_status = tree.item(selected_item)["values"][5]  # Столбец "Статус оплаты"

    if current_status == "Не оплачен":
        confirmation = messagebox.askyesno(
            "Подтверждение оплаты",
            "Вы уверены, что гость произвел оплату?"
        )
        if not confirmation:
            return

    # Определяем новый статус
    new_status = "Оплачен" if current_status == "Не оплачен" else "Не оплачен"

    try:
        query = "UPDATE Bills SET payment_status = ? WHERE bill_id = ?"
        execute_query(DB_PATH, query, (new_status, bill_id))
        messagebox.showinfo("Успех", f"Статус оплаты изменен на '{new_status}'!")
        load_bills(tree)  # Перезагрузка данных
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось изменить статус оплаты: {e}")


# Функция для экспорта данных в Excel
def export_to_excel(tree):
    # Подтверждение действия
    if not messagebox.askyesno("Подтверждение", "Вы уверены, что хотите создать отчет в Excel?"):
        return

    try:
        # Создаём новый Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Счета"

        # Получаем заголовки из дерева
        headers = [tree.heading(col)["text"] for col in tree["columns"]]
        column_widths = [tree.column(col)["width"] for col in tree["columns"]]

        # Заполняем заголовки в Excel
        for col_num, (header, width) in enumerate(zip(headers, column_widths), start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)  # Заголовки жирным
            # Настраиваем ширину столбцов Excel (переводим в условные единицы)
            ws.column_dimensions[cell.column_letter].width = width // 6  # Примерное преобразование

        # Заполняем данные из таблицы
        total_paid = 0  # Для подсчета суммы оплаченных счетов
        for row_num, item in enumerate(tree.get_children(), start=2):
            values = tree.item(item, "values")
            for col_num, value in enumerate(values, start=1):
                ws.cell(row=row_num, column=col_num, value=value)
            # Если счет оплачен, добавляем его сумму к итогу
            if values[-1] == "Оплачен":
                total_paid += float(values[4])

        # Добавляем итоговую строку
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1, value="ИТОГ:")
        ws.cell(row=last_row, column=5, value=total_paid).font = Font(bold=True)

        # Сохраняем файл
        file_path = "Счета_отчет.xlsx"
        wb.save(file_path)
        messagebox.showinfo("Успех", f"Отчет успешно создан: {file_path}")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось создать отчет: {e}")


# Интерфейс управления счетами
def open_bill_management(root):
    root.withdraw()

    global bill_window
    bill_window = tk.Toplevel()
    setup_window(bill_window, "Управление счетами", 1000, 600)

    tk.Label(bill_window, text="Управление счетами", font=("Arial", 20, "bold")).pack(pady=20)

    def on_close():
        bill_window.destroy()
        root.deiconify()

    bill_window.protocol("WM_DELETE_WINDOW", on_close)

    # Определяем столбцы
    columns = ("ID счета", "ID брони", "Гость", "Комната", "Сумма, руб.", "Статус оплаты")
    tree = ttk.Treeview(bill_window, columns=columns, show="headings", height=15)

    # Настраиваем ширину столбцов
    tree.column("ID счета", width=100)
    tree.column("ID брони", width=100)
    tree.column("Гость", width=200)
    tree.column("Комната", width=200)
    tree.column("Сумма, руб.", width=100)
    tree.column("Статус оплаты", width=150)

    # Настраиваем заголовки
    for col in columns:
        tree.heading(col, text=col)

    # Добавляем вертикальный скроллбар
    scrollbar = ttk.Scrollbar(bill_window, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(pady=20, fill=tk.BOTH, expand=True)

    # Кнопки управления
    button_frame = tk.Frame(bill_window)
    button_frame.pack(pady=10)

    tk.Button(button_frame, text="Оплачен", font=("Arial", 14),
              command=lambda: change_payment_status(tree)).grid(row=0, column=0, padx=10)
    tk.Button(button_frame, text="Удалить счет", font=("Arial", 14),
              command=lambda: delete_bill(tree)).grid(row=0, column=1, padx=10)
    tk.Button(button_frame, text="Создать отчет", font=("Arial", 14),
              command=lambda: export_to_excel(tree)).grid(row=0, column=2, padx=10)

    # Поле поиска
    search_frame = tk.Frame(bill_window)
    search_frame.pack(pady=10)

    tk.Label(search_frame, text="Поиск:", font=("Arial", 14)).grid(row=0, column=0, padx=10)

    search_entry = tk.Entry(search_frame, font=("Arial", 14))
    search_entry.grid(row=0, column=1, padx=10)

    def on_search_change(event):
        search_query = search_entry.get()
        load_bills(tree, search_query)

    search_entry.bind("<KeyRelease>", on_search_change)

    # Загружаем счета в таблицу
    load_bills(tree)
